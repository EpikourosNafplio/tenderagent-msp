"""
TenderNed Dataset Importer

Leest de TenderNed openbare dataset (Excel of JSON) en laadt deze in een
SQLite database die de TenderAgent gebruikt voor:
  - Gunningshistorie per opdrachtgever
  - Herhalingspatronen (verwachte heraanbestedingen)
  - Vooraankondigingen en marktconsultaties

Gebruik:
  python import_dataset.py /pad/naar/tenderned_dataset.xlsx
  python import_dataset.py /pad/naar/2025.json

De dataset is te downloaden op:
  https://www.tenderned.nl/cms/nl/aanbesteden-in-cijfers/datasets-aanbestedingen

Output: data/tenderned_historie.db (SQLite)
"""

import json
import sqlite3
import sys
import os
from pathlib import Path
from datetime import datetime

# ICT-gerelateerde CPV-codes (2-digit prefix matching)
ICT_CPV_PREFIXES = [
    "72",  # IT-diensten
    "48",  # Software
    "30",  # Computeruitrusting (30.2x)
    "64",  # Telecommunicatie (64.2x)
    "50",  # Reparatie computers (50.3x)
    "51",  # Installatie kantooruitrusting (51.6x)
]

# Meer specifieke 4-digit prefixes voor bredere categorieeen
ICT_CPV_SPECIFIC = [
    "3020", "3021", "3023",  # Computeruitrusting
    "6420", "6421",          # Telecommunicatie
    "5030",                  # Reparatie computers
    "5160",                  # Installatie kantooruitrusting
]

DB_PATH = Path(__file__).parent / "data" / "tenderned_historie.db"


def is_ict_related(cpv_codes_str: str, beschrijving: str = "") -> bool:
    """Check of een publicatie ICT-gerelateerd is op basis van CPV-codes en beschrijving."""
    if not cpv_codes_str:
        # Fallback: check beschrijving
        beschrijving_lower = beschrijving.lower()
        ict_keywords = [
            "ict", "it-dienst", "software", "hosting", "cloud",
            "werkplek", "datacenter", "informatiebeveiliging", "cybersecurity",
            "netwerk", "firewall", "applicatie", "servicedesk", "helpdesk",
            "server", "storage", "backup", "digitalisering",
        ]
        return any(kw in beschrijving_lower for kw in ict_keywords)

    cpv_str = cpv_codes_str.replace(" ", "")

    # Check op 72xxx en 48xxx (altijd ICT)
    if "72" in cpv_str[:20] or "48" in cpv_str[:20]:
        return True

    # Check alle CPV-codes
    codes = [c.strip() for c in cpv_codes_str.split(",")]
    for code in codes:
        code_clean = code.strip()
        if code_clean.startswith("72") or code_clean.startswith("48"):
            return True
        for prefix in ICT_CPV_SPECIFIC:
            if code_clean.startswith(prefix):
                return True

    return False


def create_database():
    """Maak de SQLite database en tabellen aan."""
    os.makedirs(DB_PATH.parent, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")

    conn.executescript("""
        DROP TABLE IF EXISTS percelen;
        DROP TABLE IF EXISTS gunningen;

        CREATE TABLE gunningen (
            publicatie_id TEXT PRIMARY KEY,
            tenderned_kenmerk TEXT,
            publicatiedatum TEXT,
            publicatie_soort TEXT,
            aanbestedende_dienst TEXT,
            officiele_naam TEXT,
            beschrijving TEXT,
            type_opdracht TEXT,
            procedure_type TEXT,
            nationaal_europees TEXT,
            cpv_codes TEXT,
            is_ict INTEGER DEFAULT 0
        );

        CREATE TABLE percelen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            publicatie_id TEXT,
            perceel_id TEXT,
            naam_perceel TEXT,
            datum_gunning TEXT,
            datum_winnaar_gekozen TEXT,
            aantal_inschrijvingen INTEGER,
            aantal_elektronisch INTEGER,
            gegunde_ondernemer TEXT,
            gegunde_adres TEXT,
            gegunde_plaats TEXT,
            gegunde_postcode TEXT,
            gegunde_land TEXT,
            gegunde_website TEXT,
            geraamde_waarde REAL,
            geraamde_btw_percentage REAL,
            definitieve_waarde REAL,
            definitieve_valuta TEXT,
            FOREIGN KEY (publicatie_id) REFERENCES gunningen(publicatie_id)
        );

        CREATE INDEX idx_gunningen_dienst ON gunningen(aanbestedende_dienst);
        CREATE INDEX idx_gunningen_soort ON gunningen(publicatie_soort);
        CREATE INDEX idx_gunningen_ict ON gunningen(is_ict);
        CREATE INDEX idx_gunningen_datum ON gunningen(publicatiedatum);
        CREATE INDEX idx_percelen_pubid ON percelen(publicatie_id);
        CREATE INDEX idx_percelen_gunning ON percelen(datum_gunning);
        CREATE INDEX idx_percelen_ondernemer ON percelen(gegunde_ondernemer);
    """)

    conn.commit()
    return conn


def import_excel(filepath: str, conn: sqlite3.Connection) -> tuple[int, int]:
    """Importeer TenderNed Excel dataset."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl niet geinstalleerd. Draai: pip install openpyxl")
        sys.exit(1)

    print(f"Openen van {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    total_rows = 0
    ict_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nVerwerken sheet: {sheet_name}")

        # Lees headers
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row)]
            break

        if not headers:
            continue

        # Map kolomnamen (TenderNed dataset heeft wisselende kolomnamen)
        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.lower().strip()
            if "publicatie" in h_lower and "id" in h_lower:
                col_map["publicatie_id"] = i
            elif "tenderned" in h_lower and "kenmerk" in h_lower:
                col_map["tenderned_kenmerk"] = i
            elif "publicatiedatum" in h_lower:
                col_map["publicatiedatum"] = i
            elif "publicatie" in h_lower and "soort" in h_lower:
                col_map["publicatie_soort"] = i
            elif ("aanbestedende" in h_lower and "dienst" in h_lower) or "naam aanbestedende dienst" in h_lower:
                if "officiel" not in h_lower:
                    col_map["aanbestedende_dienst"] = i
                else:
                    col_map["officiele_naam"] = i
            elif "offici" in h_lower and "naam" in h_lower:
                col_map["officiele_naam"] = i
            elif "korte beschrijving" in h_lower or ("beschrijving" in h_lower and "kort" in h_lower):
                col_map["beschrijving"] = i
            elif "type opdracht" in h_lower or ("type" in h_lower and "opdracht" in h_lower):
                col_map["type_opdracht"] = i
            elif "procedure" in h_lower and "type" not in h_lower:
                col_map["procedure_type"] = i
            elif "nationaal" in h_lower or "europees" in h_lower:
                col_map["nationaal_europees"] = i
            elif "cpv" in h_lower:
                col_map["cpv_codes"] = i
            # Perceel-velden
            elif "id perceel" in h_lower or "perceel" in h_lower and "id" in h_lower:
                col_map["perceel_id"] = i
            elif "naam perceel" in h_lower:
                col_map["naam_perceel"] = i
            elif "datum gunning" in h_lower or "gunningsdatum" in h_lower:
                if "winnaar" not in h_lower:
                    col_map["datum_gunning"] = i
                else:
                    col_map["datum_winnaar_gekozen"] = i
            elif "winnaar" in h_lower and "datum" in h_lower:
                col_map["datum_winnaar_gekozen"] = i
            elif "aantal inschrijvingen" in h_lower:
                if "elektronisch" in h_lower:
                    col_map["aantal_elektronisch"] = i
                else:
                    col_map["aantal_inschrijvingen"] = i
            elif "naam" in h_lower and ("gegund" in h_lower or "ondernemer" in h_lower or "winnaar" in h_lower):
                col_map["gegunde_ondernemer"] = i
            elif "adres" in h_lower and ("gegund" in h_lower or "ondernemer" in h_lower):
                col_map["gegunde_adres"] = i
            elif "plaats" in h_lower and ("gegund" in h_lower or "ondernemer" in h_lower):
                col_map["gegunde_plaats"] = i
            elif "postcode" in h_lower and ("gegund" in h_lower or "ondernemer" in h_lower):
                col_map["gegunde_postcode"] = i
            elif "land" in h_lower and ("gegund" in h_lower or "ondernemer" in h_lower):
                col_map["gegunde_land"] = i
            elif "website" in h_lower:
                col_map["gegunde_website"] = i
            elif "geraamde" in h_lower and "waarde" in h_lower and "btw" not in h_lower:
                col_map["geraamde_waarde"] = i
            elif "geraamde" in h_lower and "btw" in h_lower:
                col_map["geraamde_btw"] = i
            elif "definitieve" in h_lower and "waarde" in h_lower and "valuta" not in h_lower:
                col_map["definitieve_waarde"] = i
            elif "definitieve" in h_lower and "valuta" in h_lower:
                col_map["definitieve_valuta"] = i

        print(f"  Kolommen gevonden: {list(col_map.keys())}")

        def get_val(row_data, key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row_data):
                val = row_data[idx]
                if val is not None:
                    return str(val).strip()
            return ""

        def get_float(row_data, key):
            val = get_val(row_data, key)
            if val:
                try:
                    return float(val.replace(",", ".").replace("€", "").replace(" ", ""))
                except ValueError:
                    return None
            return None

        def get_int(row_data, key):
            val = get_val(row_data, key)
            if val:
                try:
                    return int(float(val))
                except ValueError:
                    return None
            return None

        batch_gunningen = []
        batch_percelen = []
        seen_pub_ids = set()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1

            pub_id = get_val(row, "publicatie_id")
            if not pub_id:
                continue

            cpv = get_val(row, "cpv_codes")
            beschrijving = get_val(row, "beschrijving")
            is_ict = is_ict_related(cpv, beschrijving)

            # Gunningen tabel (deduplicate op publicatie_id)
            if pub_id not in seen_pub_ids:
                seen_pub_ids.add(pub_id)
                batch_gunningen.append((
                    pub_id,
                    get_val(row, "tenderned_kenmerk"),
                    get_val(row, "publicatiedatum"),
                    get_val(row, "publicatie_soort"),
                    get_val(row, "aanbestedende_dienst"),
                    get_val(row, "officiele_naam"),
                    beschrijving,
                    get_val(row, "type_opdracht"),
                    get_val(row, "procedure_type"),
                    get_val(row, "nationaal_europees"),
                    cpv,
                    1 if is_ict else 0,
                ))
                if is_ict:
                    ict_rows += 1

            # Percelen tabel (als er gunningsinformatie is)
            gegunde = get_val(row, "gegunde_ondernemer")
            datum_g = get_val(row, "datum_gunning")
            if gegunde or datum_g:
                batch_percelen.append((
                    pub_id,
                    get_val(row, "perceel_id"),
                    get_val(row, "naam_perceel"),
                    datum_g,
                    get_val(row, "datum_winnaar_gekozen"),
                    get_int(row, "aantal_inschrijvingen"),
                    get_int(row, "aantal_elektronisch"),
                    gegunde,
                    get_val(row, "gegunde_adres"),
                    get_val(row, "gegunde_plaats"),
                    get_val(row, "gegunde_postcode"),
                    get_val(row, "gegunde_land"),
                    get_val(row, "gegunde_website"),
                    get_float(row, "geraamde_waarde"),
                    get_float(row, "geraamde_btw"),
                    get_float(row, "definitieve_waarde"),
                    get_val(row, "definitieve_valuta"),
                ))

            # Batch insert elke 5000 rijen
            if len(batch_gunningen) >= 5000:
                conn.executemany(
                    "INSERT OR IGNORE INTO gunningen VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    batch_gunningen
                )
                if batch_percelen:
                    conn.executemany(
                        "INSERT INTO percelen (publicatie_id, perceel_id, naam_perceel, "
                        "datum_gunning, datum_winnaar_gekozen, aantal_inschrijvingen, "
                        "aantal_elektronisch, gegunde_ondernemer, gegunde_adres, "
                        "gegunde_plaats, gegunde_postcode, gegunde_land, gegunde_website, "
                        "geraamde_waarde, geraamde_btw_percentage, definitieve_waarde, "
                        "definitieve_valuta) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        batch_percelen
                    )
                conn.commit()
                batch_gunningen = []
                batch_percelen = []
                print(f"  {total_rows} rijen verwerkt, {ict_rows} ICT-gerelateerd...")

        # Laatste batch
        if batch_gunningen:
            conn.executemany(
                "INSERT OR IGNORE INTO gunningen VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                batch_gunningen
            )
        if batch_percelen:
            conn.executemany(
                "INSERT INTO percelen (publicatie_id, perceel_id, naam_perceel, "
                "datum_gunning, datum_winnaar_gekozen, aantal_inschrijvingen, "
                "aantal_elektronisch, gegunde_ondernemer, gegunde_adres, "
                "gegunde_plaats, gegunde_postcode, gegunde_land, gegunde_website, "
                "geraamde_waarde, geraamde_btw_percentage, definitieve_waarde, "
                "definitieve_valuta) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                batch_percelen
            )
        conn.commit()

    wb.close()
    return total_rows, ict_rows


def import_json(filepath: str, conn: sqlite3.Connection) -> tuple[int, int]:
    """Importeer TenderNed JSON dataset (per jaar)."""
    print(f"Openen van {filepath}...")

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        records = data.get("records", data.get("data", [data]))
    elif isinstance(data, list):
        records = data
    else:
        print(f"Onverwacht JSON-formaat")
        return 0, 0

    total = 0
    ict = 0

    for record in records:
        total += 1
        pub_id = str(record.get("ID publicatie", record.get("publicatie_id", "")))
        if not pub_id:
            continue

        cpv = str(record.get("CPV-codes", record.get("cpv_codes", "")))
        beschrijving = str(record.get("Korte beschrijving aanbesteding", record.get("beschrijving", "")))
        is_ict_flag = is_ict_related(cpv, beschrijving)
        if is_ict_flag:
            ict += 1

        conn.execute(
            "INSERT OR IGNORE INTO gunningen VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                pub_id,
                str(record.get("TenderNed kenmerk", "")),
                str(record.get("Publicatiedatum", "")),
                str(record.get("Publicatie soort", "")),
                str(record.get("Naam aanbestedende dienst", "")),
                str(record.get("Officiele naam", "")),
                beschrijving,
                str(record.get("Type opdracht", "")),
                str(record.get("Procedure", "")),
                str(record.get("Nationaal/Europees", "")),
                cpv,
                1 if is_ict_flag else 0,
            )
        )

        # Perceel-info
        gegunde = str(record.get("Naam gegunde ondernemer", ""))
        datum_g = str(record.get("Datum gunning", ""))
        if gegunde or datum_g:
            def safe_float(val):
                if val:
                    try:
                        return float(str(val).replace(",", ".").replace(" ", ""))
                    except ValueError:
                        pass
                return None

            def safe_int(val):
                if val:
                    try:
                        return int(float(str(val)))
                    except ValueError:
                        pass
                return None

            conn.execute(
                "INSERT INTO percelen (publicatie_id, perceel_id, naam_perceel, "
                "datum_gunning, datum_winnaar_gekozen, aantal_inschrijvingen, "
                "aantal_elektronisch, gegunde_ondernemer, gegunde_adres, "
                "gegunde_plaats, gegunde_postcode, gegunde_land, gegunde_website, "
                "geraamde_waarde, geraamde_btw_percentage, definitieve_waarde, "
                "definitieve_valuta) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    pub_id,
                    str(record.get("ID perceel", "")),
                    str(record.get("Naam perceel", "")),
                    datum_g,
                    str(record.get("Datum wanneer winnaar is gekozen", "")),
                    safe_int(record.get("Aantal inschrijvingen")),
                    safe_int(record.get("Aantal elektronisch ingediende inschrijvingen")),
                    gegunde,
                    str(record.get("Adres gegunde ondernemer", "")),
                    str(record.get("Plaats gegunde ondernemer", "")),
                    str(record.get("Postcode gegunde ondernemer", "")),
                    str(record.get("Land gegunde ondernemer", "")),
                    str(record.get("Website gegunde ondernemer", "")),
                    safe_float(record.get("Geraamde waarde")),
                    safe_float(record.get("BTW-percentage geraamde waarde")),
                    safe_float(record.get("Definitieve waarde")),
                    str(record.get("Valuta definitieve waarde", "")),
                )
            )

        if total % 10000 == 0:
            conn.commit()
            print(f"  {total} records verwerkt, {ict} ICT-gerelateerd...")

    conn.commit()
    return total, ict


def main():
    if len(sys.argv) < 2:
        print("Gebruik: python import_dataset.py <pad_naar_dataset>")
        print("")
        print("Ondersteunde formaten:")
        print("  .xlsx  - TenderNed Excel dataset (2016-2025)")
        print("  .json  - TenderNed JSON dataset (per jaar)")
        print("")
        print("Meerdere bestanden: python import_dataset.py 2024.json 2025.json")
        print("")
        print("Download: https://www.tenderned.nl/cms/nl/aanbesteden-in-cijfers/datasets-aanbestedingen")
        sys.exit(1)

    print("=" * 60)
    print("TenderNed Dataset Importer")
    print("=" * 60)

    conn = create_database()
    print(f"Database aangemaakt: {DB_PATH}")

    total_all = 0
    ict_all = 0

    for filepath in sys.argv[1:]:
        if not os.path.exists(filepath):
            print(f"Bestand niet gevonden: {filepath}")
            continue

        ext = Path(filepath).suffix.lower()
        if ext == ".xlsx":
            total, ict = import_excel(filepath, conn)
        elif ext == ".json":
            total, ict = import_json(filepath, conn)
        else:
            print(f"Onbekend formaat: {ext}. Ondersteund: .xlsx, .json")
            continue

        total_all += total
        ict_all += ict
        print(f"  Klaar: {total} rijen, waarvan {ict} ICT-gerelateerd")

    # Statistieken
    print("\n" + "=" * 60)
    print("RESULTAAT")
    print("=" * 60)
    gun_count = conn.execute("SELECT COUNT(*) FROM gunningen").fetchone()[0]
    ict_count = conn.execute("SELECT COUNT(*) FROM gunningen WHERE is_ict = 1").fetchone()[0]
    perc_count = conn.execute("SELECT COUNT(*) FROM percelen").fetchone()[0]

    print(f"Totaal publicaties:        {gun_count}")
    print(f"Waarvan ICT-gerelateerd:   {ict_count}")
    print(f"Percelen met gunningsinfo: {perc_count}")

    # Top aanbestedende diensten (ICT)
    print("\nTop 10 ICT-aanbestedende diensten:")
    top = conn.execute(
        """SELECT aanbestedende_dienst, COUNT(*) as n
           FROM gunningen WHERE is_ict = 1
           GROUP BY aanbestedende_dienst ORDER BY n DESC LIMIT 10"""
    ).fetchall()
    for row in top:
        print(f"  {row[1]:4d}x  {row[0]}")

    # Publicatiesoorten
    print("\nPublicatiesoorten (ICT):")
    soorten = conn.execute(
        """SELECT publicatie_soort, COUNT(*) as n
           FROM gunningen WHERE is_ict = 1
           GROUP BY publicatie_soort ORDER BY n DESC"""
    ).fetchall()
    for row in soorten:
        print(f"  {row[1]:4d}x  {row[0]}")

    print(f"\nDatabase: {DB_PATH}")
    print(f"Grootte:  {os.path.getsize(DB_PATH) / 1_000_000:.1f} MB")
    print("\nKlaar! Start de TenderAgent opnieuw om de data te gebruiken.")

    conn.close()


if __name__ == "__main__":
    main()
